#!/usr/bin/env python3
"""
生成财务画像 Excel 模板。

用法：
  python3 generate_forms.py [输出路径]
  默认输出到当前目录：财务画像模板.xlsx
"""

import sys
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("错误：需要 openpyxl。安装：pip3 install openpyxl")
    sys.exit(1)

# ── 样式定义 ──

HEADER_FONT = Font(name="PingFang SC", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
SECTION_FONT = Font(name="PingFang SC", size=11, bold=True, color="2B579A")
INSTRUCTION_FONT = Font(name="PingFang SC", size=9, color="888888")
EXAMPLE_FONT = Font(name="PingFang SC", size=10, color="AAAAAA")
NORMAL_FONT = Font(name="PingFang SC", size=10)
TITLE_FONT = Font(name="PingFang SC", size=13, bold=True, color="2B579A")

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

INPUT_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
SECTION_FILL = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")


def _style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_input_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    cell.font = NORMAL_FONT
    cell.alignment = LEFT_WRAP


def _style_section_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER


def _example_row(ws, row, cols, values):
    """示例行，灰色字体"""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = val
        cell.font = EXAMPLE_FONT
        cell.border = THIN_BORDER


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════
# Sheet 1: 现金快照
# ═══════════════════════════════════════════════════════════════

def _build_cash_snapshot(ws):
    ws.title = "现金快照"
    ws.sheet_properties.tabColor = "2B579A"
    ws.freeze_panes = "A3"

    _set_col_widths(ws, [22, 20, 12, 48])

    # 标题行
    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1, value="现金快照（最少填写 ☆）").font = TITLE_FONT

    # 说明
    ws.merge_cells("A2:D2")
    ws.cell(row=2, column=1, value="只需填 3 个数字，即可获得即时财务诊断。金额填数字（元），也可以写区间如 1-1.5万。").font = INSTRUCTION_FONT

    # 表头
    headers = ["项目", "填写（元）", "必填", "说明"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header_row(ws, 3, 4)

    # 数据行
    fields = [
        ("月收入（税后）", None, "是", "工资、奖金、副业等每月到手收入"),
        ("月支出", None, "是", "含房租/房贷、餐饮、交通、购物等所有月度支出"),
        ("可动用存款", None, "否", "现金、活期存款、货币基金等可立刻取出的钱"),
    ]

    for i, (label, _, required, desc) in enumerate(fields):
        row = 4 + i
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = LEFT_WRAP
        _style_input_cell(ws, row, 2)
        ws.cell(row=row, column=3, value=required).font = NORMAL_FONT
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=4, value=desc).font = INSTRUCTION_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER

    # 示例行
    _example_row(ws, 7, 4, ["（示例）", "15000", "是", "月入15000元"])
    _example_row(ws, 8, 4, ["（示例）", "8000", "是", "月支出8000元"])
    _example_row(ws, 9, 4, ["（示例）", "50000", "否", "存款5万元"])

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28


# ═══════════════════════════════════════════════════════════════
# Sheet 2: 基本信息
# ═══════════════════════════════════════════════════════════════

def _build_basic_info(ws):
    ws.title = "基本信息"
    ws.sheet_properties.tabColor = "4CAF50"
    ws.freeze_panes = "A3"

    _set_col_widths(ws, [22, 24, 12, 48])

    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1, value="基本信息").font = TITLE_FONT

    ws.merge_cells("A2:D2")
    ws.cell(row=2, column=1, value="基础信息，帮助判断你的生活阶段和城市消费水平。").font = INSTRUCTION_FONT

    headers = ["项目", "填写", "必填", "说明 / 选项"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header_row(ws, 3, 4)

    fields = [
        ("年龄段", None, "是", "25以下 / 25-30 / 30-35 / 35-40 / 40-50 / 50以上"),
        ("所在城市", None, "是", "如 北京、上海、杭州、深圳"),
        ("职业", None, "否", "如 程序员、教师、自由职业"),
        ("家庭结构", None, "否", "单身 / 已婚无娃 / 有娃 / 需赡养父母"),
    ]

    for i, (label, _, required, desc) in enumerate(fields):
        row = 4 + i
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = LEFT_WRAP
        _style_input_cell(ws, row, 2)
        ws.cell(row=row, column=3, value=required).font = NORMAL_FONT
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=4, value=desc).font = INSTRUCTION_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER

    _example_row(ws, 8, 4, ["（示例）", "30-35", "是", "30-35岁"])
    _example_row(ws, 9, 4, ["（示例）", "北京", "是", "一线城市"])

    ws.row_dimensions[1].height = 22


# ═══════════════════════════════════════════════════════════════
# Sheet 3: 资产负债表
# ═══════════════════════════════════════════════════════════════

def _build_balance_sheet(ws):
    ws.title = "资产负债表"
    ws.sheet_properties.tabColor = "FF9800"
    ws.freeze_panes = "A4"

    _set_col_widths(ws, [20, 18, 14, 14, 16, 16, 14])

    ws.merge_cells("A1:G1")
    ws.cell(row=1, column=1, value="资产负债表").font = TITLE_FONT

    ws.merge_cells("A2:G2")
    ws.cell(row=2, column=1,
        value="填写方式二选一：① 在「分类汇总」列直接填各类总额；② 在下方「产品明细」填各平台的具体产品，系统自动归池。"
    ).font = INSTRUCTION_FONT

    # ── 表头 ──
    headers = ["类别", "分类汇总（元）", "产品名称", "所在平台", "金额（元）", "持有收益（元）", "收益率"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header_row(ws, 3, 7)

    row = 4

    # ── 资产部分 ──
    _write_balance_section(ws, row, "资产", [
        ("💵 现金池", "现金、活期存款、货币基金（余额宝、零钱通等）", 3,
         [("", "", "", "", "", "", ""), ("", "", "余额宝", "支付宝", "30000", "", "")]),
        ("🎯 目标池", "债券基金、银行理财、大额存单、定期存款", 3,
         [("", "", "", "", "", "", ""), ("", "", "招商产业债券C", "天天基金", "50000", "1200", "0.024")]),
        ("🥚 金鹅池", "偏股型基金、股票、指数基金、混合基金、QDII", 3,
         [("", "", "", "", "", "", ""), ("", "", "沪深300ETF联接", "支付宝", "100000", "5000", "0.053")]),
    ])

    # ── 负债部分 ──
    row = ws.max_row + 1
    _write_balance_section(ws, row, "负债", [
        ("📅 短期负债", "信用卡、花呗、借呗、消费贷等1年内到期的负债", 3,
         [("", "", "", "", "", "", ""), ("", "", "信用卡", "招商银行", "5000", "", "")]),
        ("🏠 长期负债", "房贷、车贷等1年以上长期贷款", 3,
         [("", "", "", "", "", "", ""), ("", "", "房贷", "工商银行", "800000", "", "")]),
    ])

    # ── 净资产行 ──
    row = ws.max_row + 2
    ws.merge_cells(f"A{row}:B{row}")
    ws.cell(row=row, column=1, value="📊 净资产 = 总资产 − 总负债").font = Font(name="PingFang SC", size=10, bold=True, color="2B579A")
    ws.merge_cells(f"C{row}:G{row}")
    ws.cell(row=row, column=3, value="自动计算，无需手动填写").font = INSTRUCTION_FONT

    _set_col_widths(ws, [20, 18, 14, 14, 16, 16, 14])


def _write_balance_section(ws, start_row, section_label, categories):
    """
    写入一组资产负债分类。

    categories: [(label, desc, data_rows, examples), ...]
    """
    row = start_row

    # 区块标题
    ws.merge_cells(f"A{row}:G{row}")
    ws.cell(row=row, column=1, value=section_label).font = Font(name="PingFang SC", size=11, bold=True, color="FFFFFF")
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = PatternFill(start_color="555555", end_color="555555", fill_type="solid")
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1

    for cat_label, cat_desc, detail_rows, examples in categories:
        # 分类名称行
        ws.merge_cells(f"A{row}:G{row}")
        ws.cell(row=row, column=1, value=f"{cat_label} — {cat_desc}").font = SECTION_FONT
        _style_section_row(ws, row, 7)
        row += 1

        # 汇总行
        ws.cell(row=row, column=1, value="  ↳ 汇总").font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        _style_input_cell(ws, row, 2)
        ws.merge_cells(f"C{row}:G{row}")
        ws.cell(row=row, column=3, value="← 直接填各类总额（方式①），或留空、在下方填产品明细").font = INSTRUCTION_FONT
        row += 1

        # 明细子表头
        detail_start = row
        for col, h in enumerate(["", "", "产品名称", "所在平台", "金额（元）", "持有收益（元）", "收益率"], 1):
            ws.cell(row=row, column=col, value=h).font = Font(name="PingFang SC", size=9, color="666666")
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

        # 明细行
        for _ in range(detail_rows):
            for col in [3, 4, 5, 6, 7]:
                _style_input_cell(ws, row, col)
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=2).border = THIN_BORDER
            row += 1

        # 示例行
        if examples:
            for ex_vals in examples:
                if any(ex_vals):  # 非空行才显示
                    _example_row(ws, row, 7, ex_vals)
                    row += 1

        row += 1  # 分类间空行

    return row


# ═══════════════════════════════════════════════════════════════
# Sheet 4: 风险与目标
# ═══════════════════════════════════════════════════════════════

def _build_risk_goals(ws):
    ws.title = "风险与目标"
    ws.sheet_properties.tabColor = "E91E63"
    ws.freeze_panes = "A3"

    _set_col_widths(ws, [22, 24, 12, 48])

    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1, value="风险偏好与财务目标").font = TITLE_FONT

    ws.merge_cells("A2:D2")
    ws.cell(row=2, column=1, value="了解你的投资风格和长远想法。不想填的可以跳过。").font = INSTRUCTION_FONT

    headers = ["项目", "填写", "必填", "说明 / 选项"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header_row(ws, 3, 4)

    fields = [
        ("风险承受能力", None, "是", "保守型（接受低波动） / 平衡型（中等波动和收益） / 进取型（接受较大波动）"),
        ("财务目标", None, "否", "如 5年内存首付、10年被动收入覆盖支出、攒够退休金"),
        ("期望达成时间", None, "否", "如 3-5年、5-10年、10年以上"),
        ("教育金需求", None, "否", "如有子女，计划储备多少教育金。无则跳过"),
    ]

    for i, (label, _, required, desc) in enumerate(fields):
        row = 4 + i
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = LEFT_WRAP
        _style_input_cell(ws, row, 2)
        ws.cell(row=row, column=3, value=required).font = NORMAL_FONT
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=4, value=desc).font = INSTRUCTION_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER

    # 风险偏好下拉验证
    risk_dv = DataValidation(
        type="list",
        formula1='"保守型,平衡型,进取型"',
        allow_blank=True,
    )
    risk_dv.error = "请选择：保守型、平衡型、进取型"
    risk_dv.errorTitle = "无效输入"
    ws.add_data_validation(risk_dv)
    risk_dv.add(ws.cell(row=4, column=2))

    _example_row(ws, 8, 4, ["（示例）", "平衡型", "是", "平衡型"])
    _example_row(ws, 9, 4, ["（示例）", "5年内存够首付100万", "否", "明确的财务目标"])

    ws.row_dimensions[1].height = 22


# ═══════════════════════════════════════════════════════════════
# 入口
# ═══════════════════════════════════════════════════════════════

DEFAULT_OUTPUT = "财务画像模板.xlsx"


def generate(output_path=None):
    if output_path is None:
        output_path = DEFAULT_OUTPUT

    wb = Workbook()

    # Sheet 1
    ws1 = wb.active
    _build_cash_snapshot(ws1)

    # Sheet 2
    ws2 = wb.create_sheet()
    _build_basic_info(ws2)

    # Sheet 3
    ws3 = wb.create_sheet()
    _build_balance_sheet(ws3)

    # Sheet 4
    ws4 = wb.create_sheet()
    _build_risk_goals(ws4)

    wb.save(output_path)
    print(f"[OK] 模板已生成: {os.path.abspath(output_path)}")
    print(f"     共 4 个 Sheet：现金快照 | 基本信息 | 资产负债表 | 风险与目标")
    print(f"     用 Excel / WPS / Numbers 打开，填写黄色区域即可。")

    return os.path.abspath(output_path)


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else None
    generate(path)

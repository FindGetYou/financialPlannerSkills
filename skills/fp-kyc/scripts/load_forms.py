#!/usr/bin/env python3
"""
加载用户填好的财务画像 Excel，输出结构化数据。

用法：
  python3 load_forms.py <Excel文件路径>

输出（JSON）：
{
    "cash_snapshot": {...},
    "basic_info": {...},
    "balance_sheet": {"assets": {...}, "liabilities": {...}, "net_worth": float},
    "risk_goals": {...},
    "profile_fields": {...},   # 可直接传给 batch_collect_from_form()
}
"""

import sys
import os
import json
import re

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "需要 openpyxl。安装：pip3 install openpyxl"}))
    sys.exit(1)

_exec_dir = os.path.dirname(os.path.abspath(__file__))
_scripts_dir = os.path.abspath(os.path.join(_exec_dir, "..", "..", "..", "scripts"))
if _scripts_dir not in sys.path and os.path.isdir(_scripts_dir):
    sys.path.insert(0, _scripts_dir)
from _path_setup import init
init()

from calc import investment_return


# ═══════════════════════════════════════════════════════════════
# 工具函数
# ═══════════════════════════════════════════════════════════════

def _parse_amount(val):
    """解析金额字符串为数字。支持：15000, 1.5万, 1-1.5万（取中值）, 空"""
    if val is None:
        return 0
    s = str(val).strip().replace(",", "").replace("，", "").replace("¥", "").replace("元", "")
    if not s:
        return 0

    # "1.5万" → 15000
    m = re.match(r"([\d.]+)\s*万", s)
    if m:
        return float(m.group(1)) * 10000

    # "1-1.5万" → 取中值 12500
    m = re.match(r"([\d.]+)\s*[-~到]\s*([\d.]+)\s*万?", s)
    if m:
        a, b = float(m.group(1)), float(m.group(2))
        if "万" in s:
            a *= 10000
            b *= 10000
        return (a + b) / 2

    # 纯数字
    try:
        return float(s)
    except ValueError:
        return 0


def _parse_rate(val):
    """解析收益率字符串为小数。支持：0.1, 10%, 10"""
    if val is None:
        return None
    s = str(val).strip().replace(",", "").replace("，", "")
    if not s:
        return None
    if s.endswith("%"):
        try:
            return float(s[:-1]) / 100
        except ValueError:
            return None
    try:
        num = float(s)
        if abs(num) > 1:
            return num / 100
        return num
    except ValueError:
        return None


def _parse_cell_text(val):
    """安全转字符串，None → ''"""
    if val is None:
        return ""
    return str(val).strip()


def _is_example_row(sheet, row):
    """判断是否为示例行（含"示例"文字 或 灰色字体）。"""
    for col in range(1, 6):
        cell = sheet.cell(row=row, column=col)
        # 文字包含"示例"
        if cell.value and "示例" in str(cell.value):
            return True
        # 字体颜色为灰色
        if cell.font and cell.font.color:
            color = str(cell.font.color.rgb) if cell.font.color.rgb else ""
            if color in ("00AAAAAA", "AAAAAA", "00999999", "999999", "00C0C0C0", "C0C0C0"):
                # 有灰色字体且该行没有写有效数据 → 示例行
                return True
    return False


# ═══════════════════════════════════════════════════════════════
# Sheet 解析器
# ═══════════════════════════════════════════════════════════════

def _load_cash_snapshot(ws):
    """解析现金快照 sheet。返回 (monthly_income, monthly_expense, liquid_savings)"""
    result = {"monthly_income": 0, "monthly_expense": 0, "liquid_savings": 0}

    for row in range(4, ws.max_row + 1):
        label = _parse_cell_text(ws.cell(row=row, column=1).value)
        val = _parse_cell_text(ws.cell(row=row, column=2).value)
        if not val:
            continue
        if _is_example_row(ws, row):
            continue

        amount = _parse_amount(val)
        if "收入" in label:
            result["monthly_income"] = amount
        elif "支出" in label:
            result["monthly_expense"] = amount
        elif "存款" in label or "动用" in label:
            result["liquid_savings"] = amount

    return result


def _load_basic_info(ws):
    """解析基本信息 sheet。返回 {age, city, career, family}"""
    mapping = {
        "年龄": "age",
        "城市": "city",
        "职业": "career",
        "家庭": "family",
    }
    result = {}

    for row in range(4, ws.max_row + 1):
        label = _parse_cell_text(ws.cell(row=row, column=1).value)
        val = _parse_cell_text(ws.cell(row=row, column=2).value)
        if not val:
            continue
        if _is_example_row(ws, row):
            continue

        for keyword, field in mapping.items():
            if keyword in label:
                result[field] = val
                break

    return result


def _load_balance_sheet(ws):
    """解析资产负债表。这是最复杂的 sheet。"""
    assets = {}
    liabilities = {}

    current_section = None  # "assets" | "liabilities"
    current_pool = None     # "现金池" | "目标池" | "金鹅池" | "短期负债" | "长期负债"

    POOL_MARKERS = {
        "现金池": ("assets", "现金池"),
        "目标池": ("assets", "目标池"),
        "金鹅池": ("assets", "金鹅池"),
        "短期负债": ("liabilities", "短期负债"),
        "长期负债": ("liabilities", "长期负债"),
    }

    for row in range(1, ws.max_row + 1):
        row_text = ""
        for col in range(1, 8):
            row_text += _parse_cell_text(ws.cell(row=row, column=col).value) + " "

        # 检测区块（精确匹配："资产"或"负债"，不含子池标题）
        col1 = str(ws.cell(row=row, column=1).value or "").strip()
        if col1 == "资产":
            current_section = "assets"
            continue
        if col1 == "负债":
            current_section = "liabilities"
            continue

        # 检测池标记
        for marker, (section, pool_name) in POOL_MARKERS.items():
            if marker in row_text:
                if current_section == section or current_section is None:
                    current_pool = pool_name
                    target = assets if section == "assets" else liabilities
                    if current_pool not in target:
                        target[current_pool] = {"total": 0, "items": []}
                break

        # 跳过非数据行
        if _is_example_row(ws, row):
            continue

        # 汇总行："↳ 汇总" → 读 column 2
        if "汇总" in row_text and current_pool:
            val = _parse_cell_text(ws.cell(row=row, column=2).value)
            if val:
                target = assets if current_section == "assets" else liabilities
                if current_pool in target:
                    target[current_pool]["total"] += _parse_amount(val)

        # 产品明细行：column 3(product), 4(platform), 5(amount), 6(profit), 7(rate)
        product = _parse_cell_text(ws.cell(row=row, column=3).value)
        platform = _parse_cell_text(ws.cell(row=row, column=4).value)
        amount_str = _parse_cell_text(ws.cell(row=row, column=5).value)
        profit_str = _parse_cell_text(ws.cell(row=row, column=6).value)
        rate_str = _parse_cell_text(ws.cell(row=row, column=7).value)

        # 跳过明细子表头行（"产品名称" 等在 column 3）
        if product in ("产品名称", "", None):
            continue

        if amount_str and current_pool:
            amount = _parse_amount(amount_str)
            if amount > 0:
                target = assets if current_section == "assets" else liabilities
                if current_pool not in target:
                    target[current_pool] = {"total": 0, "items": []}

                # 解析收益和收益率（仅对资产有意义，负债忽略）
                profit_input = _parse_amount(profit_str) if profit_str else None
                rate_input = _parse_rate(rate_str)

                calc_result = investment_return(
                    holding_amount=amount,
                    profit_amount=profit_input,
                    return_rate=rate_input,
                )

                target[current_pool]["items"].append({
                    "product": product,
                    "platform": platform,
                    "amount": amount,
                    "profit_amount": calc_result["profit_amount"],
                    "return_rate": calc_result["return_rate"],
                    "cost_basis": calc_result["cost_basis"],
                })
                target[current_pool]["total"] += amount

    # 计算净资产（如果两者都有数据）
    total_assets = sum(v["total"] for v in assets.values())
    total_liabilities = sum(v["total"] for v in liabilities.values())

    return {
        "assets": assets,
        "liabilities": liabilities,
        "total_assets": total_assets,
        "total_liabilities": total_liabilities,
        "net_worth": total_assets - total_liabilities,
    }


def _load_risk_goals(ws):
    """解析风险与目标 sheet。"""
    mapping = {
        "风险": "risk_tolerance",
        "目标": "goal_financial",
        "时间": "goal_timeline",
        "教育": "goal_education",
    }
    result = {}

    for row in range(4, ws.max_row + 1):
        label = _parse_cell_text(ws.cell(row=row, column=1).value)
        val = _parse_cell_text(ws.cell(row=row, column=2).value)
        if not val:
            continue
        if _is_example_row(ws, row):
            continue

        for keyword, field in mapping.items():
            if keyword in label:
                result[field] = val
                break

    # 风险承受能力标准化
    risk_map = {"保守型": "low", "平衡型": "medium", "进取型": "high",
                "保守": "low", "平衡": "medium", "进取": "high"}
    if result.get("risk_tolerance"):
        raw = result["risk_tolerance"]
        result["risk_tolerance"] = risk_map.get(raw, raw)

    return result


# ═══════════════════════════════════════════════════════════════
# 产品自动分类器（当用户填了具体产品但没填分类汇总时使用）
# ═══════════════════════════════════════════════════════════════

CASH_POOL_KEYWORDS = [
    "货币", "余额宝", "零钱通", "活期", "现金", "存款",
    "mmf", "money market", "朝朝宝", "天天利",
]

TARGET_POOL_KEYWORDS = [
    "债券", "债基", "债", "理财", "银行", "大额存单", "定期",
    "固收", "纯债", "信用债", "利率债", "可转债",
    "bond", "fixed income",
]

GROWTH_POOL_KEYWORDS = [
    "股票", "指数", "etf", "混合", "偏股", "qdii",
    "行业", "主题", "成长", "价值", "沪深300", "中证500",
    "创业板", "科创", "纳斯达克", "标普", "恒生",
    "基金",  # 泛基金，排最后作为兜底
    "stock", "equity",
]


def classify_product(product_name):
    """
    根据产品名称自动归类到资产池。

    返回: "现金池" | "目标池" | "金鹅池" | None (无法判断)
    """
    name = product_name.lower().replace(" ", "")

    for kw in CASH_POOL_KEYWORDS:
        if kw.lower().replace(" ", "") in name:
            return "现金池"

    for kw in TARGET_POOL_KEYWORDS:
        if kw.lower().replace(" ", "") in name:
            return "目标池"

    for kw in GROWTH_POOL_KEYWORDS:
        if kw.lower().replace(" ", "") in name:
            return "金鹅池"

    return None


# ═══════════════════════════════════════════════════════════════
# 主入口
# ═══════════════════════════════════════════════════════════════

def load(file_path):
    """加载 Excel 文件，返回完整结构化数据。"""
    if not os.path.exists(file_path):
        return {"error": f"文件不存在: {file_path}"}

    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        return {"error": f"无法打开文件: {e}"}

    result = {
        "cash_snapshot": {},
        "basic_info": {},
        "balance_sheet": {},
        "risk_goals": {},
        "profile_fields": {},
        "warnings": [],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if sheet_name == "现金快照":
            result["cash_snapshot"] = _load_cash_snapshot(ws)
        elif sheet_name == "基本信息":
            result["basic_info"] = _load_basic_info(ws)
        elif sheet_name == "资产负债表":
            result["balance_sheet"] = _load_balance_sheet(ws)
        elif sheet_name == "风险与目标":
            result["risk_goals"] = _load_risk_goals(ws)

    # ── 构建 profile_fields（可直接传 batch_collect_from_form）──
    cs = result["cash_snapshot"]
    bi = result["basic_info"]
    bs = result["balance_sheet"]
    rg = result["risk_goals"]

    pf = {}

    # 现金快照 → profile
    if cs.get("monthly_income"):
        pf["income"] = str(int(cs["monthly_income"]))
    if cs.get("monthly_expense"):
        pf["expense"] = str(int(cs["monthly_expense"]))

    # 基本信息 → profile
    if bi.get("age"):
        pf["age"] = bi["age"]
    if bi.get("city"):
        pf["city"] = bi["city"]
    if bi.get("career"):
        pf["career"] = bi["career"]
    if bi.get("family"):
        pf["family"] = bi["family"]

    # 资产负债表 → profile
    assets = bs.get("assets", {})
    if assets:
        # 房产判断
        has_house = "否"
        house_keywords = ["房产", "房子", "住宅"]
        for pool_name, pool_data in assets.items():
            for item in pool_data.get("items", []):
                if any(kw in item.get("product", "") for kw in house_keywords):
                    has_house = "有"
                    break
        pf["has_house"] = has_house

        # 净资产
        nw = bs.get("net_worth")
        if nw is not None and nw != 0:
            pf["net_worth"] = str(int(nw))

        # 保险判断：检查是否有保险相关产品
        has_insurance = "否"
        insurance_keywords = ["保险", "重疾", "医疗", "寿险", "意外险"]
        for pool_name, pool_data in assets.items():
            for item in pool_data.get("items", []):
                if any(kw in item.get("product", "") for kw in insurance_keywords):
                    has_insurance = "有"
                    break
        pf["has_insurance"] = has_insurance

        # 负债总额
        liab = bs.get("liabilities", {})
        total_debt = sum(v["total"] for v in liab.values())
        if total_debt > 0:
            result["balance_sheet"]["_debt_total"] = total_debt

    # 风险目标 → profile
    if rg.get("risk_tolerance"):
        pf["risk_tolerance"] = rg["risk_tolerance"]
    if rg.get("goal_financial"):
        pf["goal_financial"] = rg["goal_financial"]
    if rg.get("goal_timeline"):
        pf["goal_timeline"] = rg["goal_timeline"]
    if rg.get("goal_education"):
        pf["goal_education"] = rg["goal_education"]

    result["profile_fields"] = pf

    # ── 检查数据质量 ──
    if len(pf) == 0:
        result["warnings"].append("未识别到任何填写数据，请确认 Excel 文件已填写")

    # 检查是否有产品明细但没填分类汇总
    for section_name, section_data in [("资产", assets), ("负债", bs.get("liabilities", {}))]:
        for pool_name, pool_data in section_data.items():
            if pool_data.get("items") and pool_data["total"] == 0:
                result["warnings"].append(
                    f"{section_name}-{pool_name}：有产品明细但未填分类汇总，已根据明细自动汇总为 {pool_data['total']:.0f} 元"
                )

    wb.close()
    return result


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python3 load_forms.py <Excel文件路径>")
        sys.exit(1)

    data = load(sys.argv[1])
    print(json.dumps(data, ensure_ascii=False, indent=2, default=str))
